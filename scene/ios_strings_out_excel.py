import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from utils.cache_utils import get_list_use_folder_cache, load_from_cache
from utils.file_utils import open_folder


# ─── .strings 解析 ────────────────────────────────────────────────────────────

def find_lproj_dirs(root_folder: str) -> list[tuple[str, str]]:
    """
    递归查找所有 *.lproj 目录。
    返回 [(lang_code, abs_path), ...]，lang_code 如 'en', 'zh-Hans', 'ja'。
    """
    results = []
    for dirpath, dirnames, _ in os.walk(root_folder):
        for d in sorted(dirnames):
            if d.endswith(".lproj"):
                lang_code = d[: -len(".lproj")]
                results.append((lang_code, os.path.join(dirpath, d)))
    return results


def _unescape(s: str) -> str:
    """
    还原 .strings 中的转义字符。
    保留 \\n \\t \\r 为字面字符串（不转换为真实换行/制表符），
    只还原 \\" → " 和 \\\\ → \\。
    """
    s = s.replace("\\\\", "\x00BACKSLASH\x00")
    s = s.replace('\\"', '"')
    s = s.replace("\x00BACKSLASH\x00", "\\")
    return s


def _strip_comments(content: str) -> str:
    """
    安全地去除 .strings 注释，不破坏 value 中的 // 或 /* 内容。
    策略：逐字符状态机，识别字符串内部 / 字符串外部 / 块注释 / 行注释。
    """
    result = []
    i = 0
    n = len(content)
    in_string = False  # 当前是否在双引号字符串内

    while i < n:
        c = content[i]

        if in_string:
            # 在字符串内，原样保留，处理转义
            if c == '\\' and i + 1 < n:
                result.append(c)
                result.append(content[i + 1])
                i += 2
            elif c == '"':
                result.append(c)
                in_string = False
                i += 1
            else:
                result.append(c)
                i += 1

        else:
            # 字符串外
            if c == '"':
                in_string = True
                result.append(c)
                i += 1
            elif c == '/' and i + 1 < n and content[i + 1] == '*':
                # 块注释：跳到 */
                i += 2
                while i < n - 1:
                    if content[i] == '*' and content[i + 1] == '/':
                        i += 2
                        break
                    i += 1
                result.append(' ')  # 替换为空格，防止相邻 token 粘连
            elif c == '/' and i + 1 < n and content[i + 1] == '/':
                # 行注释：跳到行尾
                while i < n and content[i] != '\n':
                    i += 1
            else:
                result.append(c)
                i += 1

    return ''.join(result)


def parse_strings_file(filepath: str) -> list[tuple[str, str]]:
    """
    解析 Apple .strings 文件，返回 [(key, value), ...]，按文件原始顺序，
    兼容 UTF-8 / UTF-16 / BOM。重复 key 只保留最后一次出现的位置和值。
    """
    content = ""
    for enc in ("utf-8-sig", "utf-8", "utf-16"):
        try:
            with open(filepath, "r", encoding=enc) as f:
                content = f.read()
            break
        except (UnicodeDecodeError, UnicodeError):
            continue

    if not content:
        print(f"  ⚠️  无法读取文件：{filepath}")
        return []

    content = _strip_comments(content)

    pattern = re.compile(r'"((?:[^"\\]|\\.)*)"\s*=\s*"((?:[^"\\]|\\.)*)"', re.DOTALL)

    # 用 dict 去重（保留最后出现的值），同时用 list 记录首次出现顺序
    seen: dict[str, int] = {}   # key -> index in ordered list
    ordered: list[tuple[str, str]] = []

    for m in pattern.finditer(content):
        key = _unescape(m.group(1))
        val = _unescape(m.group(2))
        if key in seen:
            # 重复 key：更新值，位置不变（保持首次出现的顺序）
            ordered[seen[key]] = (key, val)
        else:
            seen[key] = len(ordered)
            ordered.append((key, val))

    return ordered


def collect_all_strings(root_folder: str):
    """
    扫描 root_folder 内所有 .lproj，汇总所有 .strings 文件内容。
    key 顺序以第一个找到该 .strings 文件的语言为基准。

    返回：
      data       = { strings_filename: { key: { lang_code: value } } }
      key_order  = { strings_filename: [key, ...] }  按原文件顺序
      lang_codes = ['en', 'zh-Hans', ...]  （已排序）
      file_names = ['Localizable.strings', ...]  （已排序）
    """
    lproj_dirs = find_lproj_dirs(root_folder)
    if not lproj_dirs:
        return {}, {}, [], []

    lang_codes = sorted({lc for lc, _ in lproj_dirs})

    all_string_files: set[str] = set()
    for _, lproj_path in lproj_dirs:
        for f in os.listdir(lproj_path):
            if f.endswith(".strings"):
                all_string_files.add(f)

    data: dict[str, dict[str, dict[str, str]]] = {}
    key_order: dict[str, list[str]] = {}

    for strings_file in sorted(all_string_files):
        data[strings_file] = {}
        key_order[strings_file] = []
        for lang_code, lproj_path in lproj_dirs:
            filepath = os.path.join(lproj_path, strings_file)
            if not os.path.isfile(filepath):
                continue
            for key, value in parse_strings_file(filepath):
                if key not in data[strings_file]:
                    data[strings_file][key] = {}
                    key_order[strings_file].append(key)   # 首次出现时记录顺序
                data[strings_file][key][lang_code] = value

    return data, key_order, lang_codes, sorted(all_string_files)


# ─── Excel 写入 ───────────────────────────────────────────────────────────────

HEADER_BG  = "4472C4"   # 深蓝（表头）
LIGHT_GRAY = "F2F2F2"   # 浅灰（交替行）
WHITE      = "FFFFFF"


def _apply_style(cell, bold=False, bg=None, font_color="000000",
                 h_align="left", wrap=False):
    cell.font = Font(name="Arial", bold=bold, color=font_color)
    cell.alignment = Alignment(horizontal=h_align, vertical="center",
                               wrap_text=wrap)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)


def write_excel(data: dict, key_order: dict, lang_codes: list[str],
                file_names: list[str], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Localizations"

    # ── 表头行 ────────────────────────────────────────────────────────────────
    headers = ["File", "Key"] + lang_codes
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        _apply_style(cell, bold=True, bg=HEADER_BG, font_color=WHITE,
                     h_align="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ── 数据行（按原文件 key 顺序）────────────────────────────────────────────
    current_row = 2
    for strings_file in file_names:
        keys_dict = data.get(strings_file, {})
        ordered_keys = key_order.get(strings_file, [])
        if not ordered_keys:
            continue

        for row_num, key in enumerate(ordered_keys):
            lang_values = keys_dict[key]
            bg = LIGHT_GRAY if row_num % 2 == 0 else WHITE

            _apply_style(
                ws.cell(row=current_row, column=1, value=strings_file),
                bg=bg
            )
            _apply_style(
                ws.cell(row=current_row, column=2, value=key),
                bold=True, bg=bg
            )
            for col_offset, lang in enumerate(lang_codes, start=3):
                _apply_style(
                    ws.cell(row=current_row, column=col_offset,
                            value=lang_values.get(lang, "")),
                    bg=bg, wrap=True
                )
            current_row += 1

    # ── 列宽 ──────────────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 42
    for col_idx in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 35

    wb.save(output_path)
    print(f"✅ Excel 已生成：{output_path}")


# ─── 入口 ─────────────────────────────────────────────────────────────────────

def run_lproj_to_excel():
    print("=== .lproj 多语言 → Excel 合并工具 ===\n")

    # 1. 选择根目录（复用你的 cache_utils）
    cache_json = load_from_cache() or {}
    root_folder = get_list_use_folder_cache(cache_json, "LprojToExcel")
    if not root_folder:
        print("❌ 未选择文件夹，已取消。")
        return

    print(f"\n📂 目标文件夹：{root_folder}")

    # 2. 扫描 .lproj
    print("🔍 正在扫描 .lproj 目录...")
    data, key_order, lang_codes, file_names = collect_all_strings(root_folder)

    if not lang_codes:
        print("❌ 未在该目录下找到任何 .lproj 文件夹。")
        return

    print(f"  找到语言：{', '.join(lang_codes)}")
    print(f"  找到文件：{', '.join(file_names)}")

    # 3. 输出路径
    output_dir = os.path.join(root_folder, "output")
    os.makedirs(output_dir, exist_ok=True)
    folder_name = os.path.basename(root_folder.rstrip("/\\"))
    output_file = os.path.join(output_dir, f"{folder_name}_localizations.xlsx")

    # 4. 生成 Excel
    print("📝 正在生成 Excel...")
    write_excel(data, key_order, lang_codes, file_names, output_file)

    # 5. 打开输出目录（复用你的 file_utils）
    open_folder(output_dir)


if __name__ == "__main__":
    run_lproj_to_excel()