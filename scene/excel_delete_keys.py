# excel_delete_keys.py  —  Function 2, Script v15
#
# 功能说明：
#   读取一个 Excel 文件（.xlsx），其中：
#     • Sheet 1（第一个工作表）：多语言数据表，包含一个 "Key" 标识列
#     • Sheet 2（第二个工作表）：需要删除的 Key 列表（每个 Key 独占一行/单元格）
#   脚本会找到 Sheet 1 中 Key 列与 Sheet 2 中列出的 Key 匹配的所有数据行，
#   并将这些行从 Sheet 1 中删除，然后将结果保存为新的 Excel 文件。
#
# Excel 格式要求（Sheet 1）：
#   - 表格顶部可以包含任意数量的平台/文件夹配置行
#   - 必须存在一个单元格，其值为字符串 "Key"（不区分大小写），
#     该行作为数据区域的表头行
#   - 表头行之后的每一行：第 Key 列的值即为该行的唯一标识 Key
#
# Excel 格式要求（Sheet 2）：
#   - 每个需要删除的 Key 填写在第一列，每行一个
#   - 第一行可以是表头（如 "Key" / "待删除" 等），脚本会跳过非数据内容
#   - 空行自动忽略

import os
import shutil

import openpyxl

from localization.localization import get_localized_text
from utils.cache_utils import load_from_cache, get_list_use_folder_cache
from utils.file_utils import find_exc_file, open_folder


# ─── 核心处理逻辑 ─────────────────────────────────────────────────────────────

def start_delete_keys(excel_path: str, output_dir: str) -> None:
    """
    v15 核心函数：从多语言 Excel 中删除 Sheet 2 指定的 Key 行。

    参数：
        excel_path  — 源 Excel 文件完整路径
        output_dir  — 输出文件夹路径（清理后的文件将保存至此）
    """
    wb = openpyxl.load_workbook(excel_path)

    # ── 基本检查 ──────────────────────────────────────────────────────────────
    if len(wb.sheetnames) < 2:
        print("❌ Excel 文件需要至少 2 个 Sheet（Sheet1：多语言表，Sheet2：待删除 Key 列表）")
        return

    ws_data = wb.worksheets[0]   # Sheet 1：多语言数据表
    ws_keys = wb.worksheets[1]   # Sheet 2：待删除 Key 列表

    print(f"📄 Sheet 1 名称：{ws_data.title}（共 {ws_data.max_row} 行）")
    print(f"📄 Sheet 2 名称：{ws_keys.title}")

    # ── Step 1：从 Sheet 2 收集待删除的 Key 集合 ──────────────────────────────
    keys_to_delete: set[str] = set()

    for row in ws_keys.iter_rows(min_row=1, values_only=True):
        # 只读取第一列；空行及 None 自动跳过
        raw = row[0] if row else None
        if raw is None:
            continue
        val = str(raw).strip()
        if val:
            keys_to_delete.add(val)

    if not keys_to_delete:
        print("⚠️  Sheet 2 中没有找到任何有效的 Key，操作取消。")
        return

    print(f"📋 待删除 Key 共 {len(keys_to_delete)} 个")

    # ── Step 2：在 Sheet 1 中定位 "Key" 标识列 ────────────────────────────────
    key_col_idx: int | None = None   # 1-based 列索引
    key_header_row: int | None = None  # 1-based 行索引

    for row_idx, row in enumerate(ws_data.iter_rows(values_only=True), start=1):
        for col_idx, cell_val in enumerate(row, start=1):
            if isinstance(cell_val, str) and cell_val.strip().lower() == "key":
                key_col_idx = col_idx
                key_header_row = row_idx
                break
        if key_col_idx is not None:
            break

    if key_col_idx is None:
        print("❌ Sheet 1 中未找到值为 'Key' 的标识单元格，请检查表格格式。")
        return

    print(f"🔍 'Key' 标识位于 Sheet 1 第 {key_header_row} 行 / 第 {key_col_idx} 列")

    # ── Step 3：遍历数据行，收集需要删除的行号 ────────────────────────────────
    rows_to_delete: list[int] = []
    matched_keys: list[str] = []

    data_start_row = key_header_row + 1
    for row_idx in range(data_start_row, ws_data.max_row + 1):
        cell_val = ws_data.cell(row=row_idx, column=key_col_idx).value
        if cell_val is None:
            continue
        key_str = str(cell_val).strip()
        if key_str in keys_to_delete:
            rows_to_delete.append(row_idx)
            matched_keys.append(key_str)

    if not rows_to_delete:
        print("ℹ️  Sheet 1 中没有找到任何与 Sheet 2 匹配的 Key 行，输出文件与原文件相同。")
    else:
        print(f"🗑️  匹配到 {len(rows_to_delete)} 行，准备删除：")
        for k in matched_keys:
            print(f"    • {k}")

        # ── Step 4：逆序删除，避免行号偏移 ──────────────────────────────────
        for row_idx in reversed(rows_to_delete):
            ws_data.delete_rows(row_idx)

        print(f"✅ 已删除 {len(rows_to_delete)} 行，Sheet 1 当前剩余 {ws_data.max_row} 行")

    # ── Step 5：检查 Sheet 2 中是否有未匹配的 Key（给出提示）─────────────────
    matched_set = set(matched_keys)
    unmatched = keys_to_delete - matched_set
    if unmatched:
        print(f"⚠️  以下 {len(unmatched)} 个 Key 在 Sheet 1 中未找到对应行：")
        for k in sorted(unmatched):
            print(f"    • {k}")

    # ── Step 6：保存输出文件 ──────────────────────────────────────────────────
    filename = os.path.basename(excel_path)
    name_without_ext, ext = os.path.splitext(filename)
    output_filename = f"{name_without_ext}_cleaned{ext}"
    output_path = os.path.join(output_dir, output_filename)

    wb.save(output_path)
    print(f"\n📁 输出文件已保存：{output_path}")


# ─── 入口函数（供 main_application.py 调用）───────────────────────────────────

def run_excel_delete_keys() -> None:
    """Function 2 v15 入口：交互式选择 Excel 文件 → 删除指定 Key 行 → 输出结果。"""

    # 1. 从缓存加载上次使用的文件夹路径
    result = get_list_use_folder_cache(load_from_cache(), "ExcelDeleteKeys")
    if not result:
        print(get_localized_text("cancel_choose_tip"))
        return

    # 2. 在选定文件夹内查找 .xlsx 文件
    exc_path = find_exc_file(result, '.xlsx')
    if not exc_path:
        return

    # 3. 准备输出文件夹（若已存在则先清空）
    output_base = os.path.splitext(exc_path)[0]
    output_directory = os.path.join(output_base, "output")

    if os.path.exists(output_directory):
        try:
            shutil.rmtree(output_directory)
        except Exception as e:
            print(get_localized_text("delete_folder_failed", error=str(e)))

    os.makedirs(output_directory)

    # 4. 执行核心逻辑
    print(get_localized_text("start_processing"))
    start_delete_keys(exc_path, output_directory)
    print(get_localized_text("complete_processing"))

    # 5. 在 Finder/文件管理器中打开输出文件夹
    open_folder(output_directory)
